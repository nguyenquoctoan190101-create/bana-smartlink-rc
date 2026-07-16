"""Legacy synthetic-period helper; database writes are deliberately retired.

The original command wrote directly into a superseded schema and could turn
blank spreadsheet cells into zero.  Its pure transformation helpers remain
only for regression tests.  Use ``db/demo_synthetic_seed.sql`` or the reviewed
batch-import workflow for release-candidate data.
"""
from __future__ import annotations

import json
import random
from pathlib import Path

import openpyxl


PROJECT_ROOT = Path(__file__).resolve().parents[1]
MERGE_MAP_PATH = (
    PROJECT_ROOT / "DU_LIEU_CHINH_THUC" / "village_merge_map_CHINH_THUC.json"
)


def _load_merge_map() -> tuple[list[str], dict[str, str]]:
    """Return only officially confirmed old-to-new village mappings.

    The official source assigns only the northern part of Đông Sơn to Hòa
    Ninh.  A complete Đông Sơn workbook must therefore remain unresolved and
    this helper must fail closed until an approved split is machine-readable.
    """
    if not MERGE_MAP_PATH.exists():
        raise FileNotFoundError(MERGE_MAP_PATH)

    with MERGE_MAP_PATH.open(encoding="utf-8") as source:
        data = json.load(source)

    names_by_id = {
        str(row["id"]): str(row["ten"])
        for row in data.get("villages_moi", [])
        if isinstance(row, dict) and row.get("id") and row.get("ten")
    }
    confirmed: dict[str, str] = {}
    unresolved: list[str] = []
    for row in data.get("anh_xa_thon_cu", []):
        if not isinstance(row, dict) or not row.get("ten_thon_cu"):
            continue
        if row.get("mapping_status", "confirmed") != "confirmed" or not row.get(
            "new_village_id"
        ):
            unresolved.append(str(row["ten_thon_cu"]))
            continue
        target = names_by_id.get(str(row["new_village_id"]))
        if target is None:
            raise RuntimeError(f"Unknown target village id: {row['new_village_id']}")
        confirmed[str(row["ten_thon_cu"])] = target

    if unresolved:
        raise RuntimeError(
            "Cannot aggregate legacy data while village allocation is unresolved: "
            + ", ".join(unresolved)
        )
    if len(names_by_id) != 10 or not confirmed:
        raise RuntimeError("Official village mapping is incomplete")
    return list(names_by_id.values()), confirmed


def _read_and_aggregate_q2_data(
    excel_path: Path,
    new_villages: list[str],
    merge_map: dict[str, str],
) -> dict[str, dict[str, int]]:
    """Read the 22-village workbook and aggregate confirmed mappings only."""
    if not excel_path.exists():
        raise FileNotFoundError(excel_path)

    workbook = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    if "Tong hop" not in workbook.sheetnames:
        raise ValueError("Workbook does not contain the required 'Tong hop' sheet")
    sheet = workbook["Tong hop"]

    old_village_data: dict[str, dict[str, int]] = {}
    for row_index in range(5, 27):
        raw_name = sheet.cell(row=row_index, column=2).value
        if not raw_name:
            continue
        village_name = str(raw_name).strip()
        indicators: dict[str, int] = {}
        for indicator_index in range(1, 15):
            code = f"CT{indicator_index:02d}"
            raw_value = sheet.cell(row=row_index, column=indicator_index + 2).value
            if raw_value is None or isinstance(raw_value, bool):
                raise ValueError(
                    f"{village_name} {code} is blank; missing data is not zero"
                )
            if not isinstance(raw_value, int):
                raise ValueError(
                    f"{village_name} {code} is not a plain integer: {raw_value!r}"
                )
            if raw_value < 0:
                raise ValueError(f"{village_name} {code} must be non-negative")
            indicators[code] = raw_value
        old_village_data[village_name] = indicators

    aggregated = {
        village: {f"CT{index:02d}": 0 for index in range(1, 15)}
        for village in new_villages
    }
    for old_name, indicators in old_village_data.items():
        new_name = merge_map.get(old_name)
        if new_name is None:
            raise ValueError(f"Legacy village is not mapped: {old_name}")
        if new_name not in aggregated:
            raise ValueError(
                f"Mapped target is not in the official 10-village list: {new_name}"
            )
        for code, value in indicators.items():
            aggregated[new_name][code] += value
    return aggregated


def _perturb_and_sanitize(q2_values: dict[str, int]) -> dict[str, int]:
    """Create deterministic-rule-safe synthetic values for legacy tests only."""
    missing = [f"CT{index:02d}" for index in range(1, 15) if f"CT{index:02d}" not in q2_values]
    if missing:
        raise ValueError("Missing indicators: " + ", ".join(missing))
    if any(isinstance(value, bool) or not isinstance(value, int) or value < 0 for value in q2_values.values()):
        raise ValueError("All indicator values must be non-negative integers")

    values = {
        code: max(
            0,
            int(
                round(
                    value
                    * (
                        1
                        + random.uniform(0.05, 0.15)
                        * random.choice((-1, 1))
                    )
                )
            ),
        )
        for code, value in q2_values.items()
    }
    values["CT01"] = values["CT01"] or 100
    values["CT02"] = int(round(values["CT01"] * random.uniform(3.2, 4.0)))
    values["CT03"] = min(values["CT03"], int(values["CT01"] * 0.10))
    values["CT04"] = min(
        values["CT04"],
        int(values["CT01"] * 0.15),
        values["CT01"] - values["CT03"],
    )
    values["CT05"] = min(values["CT05"], int(values["CT02"] * 0.05))
    values["CT06"] = min(values["CT06"], int(values["CT02"] * 0.10))
    values["CT07"] = min(values["CT07"], int(values["CT02"] * 0.35))
    values["CT08"] = min(values["CT08"], int(values["CT07"] * 0.15))
    values["CT09"] = max(
        min(values["CT09"], values["CT01"]), int(values["CT01"] * 0.85)
    )
    values["CT10"] = min(
        max(values["CT10"], values["CT07"] + 10),
        int(values["CT02"] * 0.70),
    )
    values["CT11"] = max(
        min(values["CT11"], values["CT02"]), int(values["CT02"] * 0.92)
    )
    values["CT12"] = max(3, min(values["CT12"], 20))
    values["CT13"] = min(values["CT13"], values["CT02"])
    values["CT14"] = min(values["CT14"], 5)
    return values


def main() -> int:
    """Refuse the obsolete direct database writer."""
    print(
        "[BLOCKED] seed_fake_prior_period.py is retired. "
        "Use db/demo_synthetic_seed.sql or the reviewed report-import workflow."
    )
    return 2


if __name__ == "__main__":
    raise SystemExit(main())
