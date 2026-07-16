"""Read-only, PII-safe audit for the organizer-provided Excel package.

The source workbooks contain reporter contact fields.  This command deliberately
never serializes those fields: it emits only file integrity metadata, village
names, CT01-CT14 values, validation codes, and a boolean phone-format result.

Example::

    python scripts/audit_source_workbooks.py D:/path/to/package --output audit.json

The source directory may be the package root or any parent containing the
``02_Bao_cao_tung_thon`` and ``03_Tong_hop_va_theo_doi`` folders.
"""

from __future__ import annotations

import argparse
import hashlib
import json
from pathlib import Path
import sys
from typing import Any

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from services.excel_report_parser import parse_official_report_excel  # noqa: E402
from services.validator import validate_phone, validate_report  # noqa: E402
from services.village_mapper import resolve_village_mapping  # noqa: E402


MAPPING_PATH = PROJECT_ROOT / "DU_LIEU_CHINH_THUC" / "village_merge_map_CHINH_THUC.json"
EXPECTED_CODES = tuple(f"CT{index:02d}" for index in range(1, 15))
SUMMARY_FILENAME = "TONG_HOP_va_THEO_DOI_TIEN_DO.xlsx"


class SourceWorkbookAuditError(RuntimeError):
    """Raised when the source package cannot be audited deterministically."""


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def find_unique(root: Path, pattern: str) -> Path:
    matches = sorted(path for path in root.rglob(pattern) if path.is_file())
    if len(matches) != 1:
        raise SourceWorkbookAuditError(
            f"Expected exactly one {pattern!r} below the selected directory; found {len(matches)}."
        )
    return matches[0]


def find_unique_directory(root: Path, name: str) -> Path:
    matches = sorted(path for path in root.rglob(name) if path.is_dir())
    if root.name == name and root.is_dir():
        matches = [root, *matches]
    matches = list(dict.fromkeys(matches))
    if len(matches) != 1:
        raise SourceWorkbookAuditError(
            f"Expected exactly one directory {name!r} below the selected directory; "
            f"found {len(matches)}."
        )
    return matches[0]


def load_summary_rows(summary_path: Path) -> dict[str, dict[str, Any]]:
    workbook = load_workbook(summary_path, read_only=True, data_only=True)
    if "Tong hop" not in workbook.sheetnames:
        raise SourceWorkbookAuditError("The summary workbook is missing sheet 'Tong hop'.")

    worksheet = workbook["Tong hop"]
    rows: dict[str, dict[str, Any]] = {}
    for row in worksheet.iter_rows(min_row=5, values_only=True):
        village_name = row[1] if len(row) > 1 else None
        if not isinstance(village_name, str) or not village_name.startswith("Thôn "):
            continue
        values = {
            code: row[index + 2] if len(row) > index + 2 else None
            for index, code in enumerate(EXPECTED_CODES)
        }
        rows[village_name.strip()] = {
            "values": values,
            "submission_status": row[16] if len(row) > 16 else None,
        }
    return rows


def _validation_codes(values: dict[str, Any]) -> list[tuple[str, str]]:
    return sorted(
        {
            (
                str(issue.get("ct_code") or "REPORT"),
                str(issue.get("error_type") or "INVALID"),
            )
            for issue in validate_report(values)
        }
    )


def audit_package(source_root: Path) -> dict[str, Any]:
    source_root = source_root.resolve()
    if not source_root.is_dir():
        raise SourceWorkbookAuditError("The selected source package directory does not exist.")

    report_dir = find_unique_directory(source_root, "02_Bao_cao_tung_thon")
    summary_path = find_unique(source_root, SUMMARY_FILENAME)

    mapping = json.loads(MAPPING_PATH.read_text(encoding="utf-8"))
    summary_rows = load_summary_rows(summary_path)
    reports: list[dict[str, Any]] = []

    for path in sorted(report_dir.glob("*.xlsx"), key=lambda item: item.name.casefold()):
        parsed = parse_official_report_excel(path.read_bytes())
        village_name = parsed["metadata"]["village_name"]
        if not village_name:
            raise SourceWorkbookAuditError(f"{path.name} has no village name in the official header.")

        summary = summary_rows.get(village_name)
        if summary is None:
            raise SourceWorkbookAuditError(
                f"{path.name} names {village_name!r}, which is absent from the summary workbook."
            )

        differences = [
            {
                "ct_code": code,
                "source_value": parsed["values"].get(code),
                "reviewed_summary_value": summary["values"].get(code),
            }
            for code in EXPECTED_CODES
            if parsed["values"].get(code) != summary["values"].get(code)
        ]
        mapping_result = resolve_village_mapping(village_name, mapping)
        reports.append(
            {
                "filename": path.name,
                "sha256": sha256_file(path),
                "size_bytes": path.stat().st_size,
                "village_name": village_name,
                "raw_values": {code: parsed["values"].get(code) for code in EXPECTED_CODES},
                "validation_issues": [
                    {"ct_code": code, "error_type": error_type}
                    for code, error_type in _validation_codes(parsed["values"])
                ],
                "reporter_phone_format_valid": validate_phone(
                    parsed["metadata"]["reporter_phone"]
                )
                is None,
                "mapping": {
                    "status": mapping_result["mapping_status"],
                    "target_village_id": mapping_result["target_village_id"],
                    "proposed_target_village_id": mapping_result["proposed_target_village_id"],
                },
                "reviewed_summary_differences": differences,
                "summary_submission_status": summary["submission_status"],
            }
        )

    submitted_villages = {report["village_name"] for report in reports}
    missing = sorted(set(summary_rows) - submitted_villages, key=str.casefold)
    unresolved = [
        report["village_name"]
        for report in reports
        if report["mapping"]["status"] != "confirmed"
    ]

    return {
        "schema_version": 1,
        "classification": "redacted-source-audit",
        "contains_reporter_pii": False,
        "source_package_name": source_root.name,
        "summary_workbook": {
            "filename": summary_path.name,
            "sha256": sha256_file(summary_path),
            "size_bytes": summary_path.stat().st_size,
        },
        "counts": {
            "villages_in_summary": len(summary_rows),
            "report_workbooks": len(reports),
            "missing_report_workbooks": len(missing),
        },
        "missing_report_villages": missing,
        "unresolved_mapping_villages": unresolved,
        "reports": reports,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("source_root", type=Path)
    parser.add_argument("--output", type=Path)
    args = parser.parse_args()

    try:
        audit = audit_package(args.source_root)
    except SourceWorkbookAuditError as exc:
        parser.error(str(exc))

    payload = json.dumps(audit, ensure_ascii=False, indent=2) + "\n"
    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(payload, encoding="utf-8")
        print(f"Wrote redacted audit: {args.output}")
    else:
        print(payload, end="")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
