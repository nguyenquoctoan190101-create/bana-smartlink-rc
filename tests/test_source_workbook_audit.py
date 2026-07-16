from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook

from scripts import audit_source_workbooks


def _write_summary(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Tong hop"
    worksheet.cell(5, 2, "Thôn Kiểm Thử")
    for offset in range(14):
        worksheet.cell(5, offset + 3, offset + 1)
    worksheet.cell(5, 17, "Đúng hạn")
    workbook.save(path)


def test_source_audit_is_redacted_and_records_reviewed_differences(
    tmp_path: Path, monkeypatch
) -> None:
    report_dir = tmp_path / "02_Bao_cao_tung_thon"
    report_dir.mkdir()
    source_file = report_dir / "BC_T01_Thon_Kiem_Thu.xlsx"
    source_file.write_bytes(b"source-workbook-bytes")
    _write_summary(tmp_path / audit_source_workbooks.SUMMARY_FILENAME)

    mapping_path = tmp_path / "mapping.json"
    mapping_path.write_text(
        json.dumps(
            {
                "anh_xa_thon_cu": [
                    {
                        "ten_thon_cu": "Thôn Kiểm Thử",
                        "new_village_id": "kiem_thu",
                    }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    monkeypatch.setattr(audit_source_workbooks, "MAPPING_PATH", mapping_path)
    monkeypatch.setattr(
        audit_source_workbooks,
        "parse_official_report_excel",
        lambda _content: {
            "values": {f"CT{index:02d}": index for index in range(1, 15)},
            "notes": {},
            "metadata": {
                "period_name": "Kỳ kiểm thử",
                "village_name": "Thôn Kiểm Thử",
                "reporter_name": "Tên tuyệt đối không được xuất",
                "reporter_title": "Cán bộ",
                "reporter_phone": "0901234567",
                "deadline": None,
            },
        },
    )
    monkeypatch.setattr(audit_source_workbooks, "validate_report", lambda _values: [])
    monkeypatch.setattr(audit_source_workbooks, "validate_phone", lambda _phone: None)

    result = audit_source_workbooks.audit_package(tmp_path)
    serialized = json.dumps(result, ensure_ascii=False)

    assert result["counts"] == {
        "villages_in_summary": 1,
        "report_workbooks": 1,
        "missing_report_workbooks": 0,
    }
    assert result["reports"][0]["mapping"] == {
        "status": "confirmed",
        "target_village_id": "kiem_thu",
        "proposed_target_village_id": None,
    }
    assert result["reports"][0]["reviewed_summary_differences"] == []
    assert "Tên tuyệt đối không được xuất" not in serialized
    assert "0901234567" not in serialized


def test_find_unique_directory_rejects_ambiguous_packages(tmp_path: Path) -> None:
    (tmp_path / "a" / "02_Bao_cao_tung_thon").mkdir(parents=True)
    (tmp_path / "b" / "02_Bao_cao_tung_thon").mkdir(parents=True)

    try:
        audit_source_workbooks.find_unique_directory(tmp_path, "02_Bao_cao_tung_thon")
    except audit_source_workbooks.SourceWorkbookAuditError as exc:
        assert "found 2" in str(exc)
    else:
        raise AssertionError("Ambiguous source packages must fail closed")
