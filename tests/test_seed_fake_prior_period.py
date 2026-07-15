from __future__ import annotations

from pathlib import Path


from scripts.seed_fake_prior_period import (
    _perturb_and_sanitize,
    _read_and_aggregate_q2_data,
)


def test_perturb_and_sanitize_enforces_rules() -> None:
    # Set up some bad perturbed values
    bad_values = {
        "CT01": 100,
        "CT02": 50,  # too small (< 3 * CT01)
        "CT03": 120, # > CT01 ( nghèo > tổng hộ)
        "CT04": 50,  # CT03 + CT04 > CT01
        "CT05": 10,
        "CT06": 20,
        "CT07": 150, # > CT02
        "CT08": 200, # > CT07
        "CT09": 50,  # < 85% of CT01
        "CT10": 200, # > CT02
        "CT11": 300, # > CT02
        "CT12": 50,  # > 20
        "CT13": 500, # > CT02
        "CT14": 10,  # > 5
    }

    sanitized = _perturb_and_sanitize(bad_values)

    # Assert rules are satisfied
    assert 85 <= sanitized["CT01"] <= 115
    assert 3 * sanitized["CT01"] <= sanitized["CT02"] <= 4.5 * sanitized["CT01"]
    assert sanitized["CT03"] <= sanitized["CT01"] * 0.1
    assert sanitized["CT03"] + sanitized["CT04"] <= sanitized["CT01"]
    assert sanitized["CT07"] <= sanitized["CT02"]
    assert sanitized["CT08"] <= sanitized["CT07"]
    assert sanitized["CT09"] >= int(sanitized["CT01"] * 0.85)
    assert sanitized["CT10"] <= sanitized["CT02"]
    assert sanitized["CT11"] <= sanitized["CT02"]
    assert 3 <= sanitized["CT12"] <= 20
    assert sanitized["CT13"] <= sanitized["CT02"]
    assert sanitized["CT14"] <= 5


def test_read_and_aggregate_q2_data(tmp_path: Path) -> None:
    # Create fake merge map
    new_villages = ["Thôn Một", "Thôn Phước Thái"]
    merge_map = {
        "Thôn Phú Hòa 1": "Thôn Một",
        "Thôn Phú Hòa 2": "Thôn Một",
        "Thôn Phước Thái": "Thôn Phước Thái",
    }

    # Write a test workbook
    import openpyxl
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Tong hop"

    # Write headers
    sheet.cell(row=4, column=2, value="Thôn")
    for idx in range(1, 15):
        sheet.cell(row=4, column=idx + 2, value=f"CT{idx:02d}")

    # Write row 5 (Thôn Phú Hòa 1)
    sheet.cell(row=5, column=2, value="Thôn Phú Hòa 1")
    for idx in range(1, 15):
        sheet.cell(row=5, column=idx + 2, value=10)

    # Write row 6 (Thôn Phú Hòa 2)
    sheet.cell(row=6, column=2, value="Thôn Phú Hòa 2")
    for idx in range(1, 15):
        sheet.cell(row=6, column=idx + 2, value=20)

    # Write row 7 (Thôn Phước Thái)
    sheet.cell(row=7, column=2, value="Thôn Phước Thái")
    for idx in range(1, 15):
        sheet.cell(row=7, column=idx + 2, value=100)

    excel_file = tmp_path / "test.xlsx"
    wb.save(excel_file)

    aggregated = _read_and_aggregate_q2_data(excel_file, new_villages, merge_map)

    # Check aggregation:
    # Thôn Một = Thôn Phú Hòa 1 (10) + Thôn Phú Hòa 2 (20) = 30
    assert aggregated["Thôn Một"]["CT01"] == 30
    # Thôn Phước Thái = 100
    assert aggregated["Thôn Phước Thái"]["CT01"] == 100
