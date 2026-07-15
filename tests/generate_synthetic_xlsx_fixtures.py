"""Generate the sanitized Excel golden set used by hermetic CI tests.

The workbooks intentionally contain synthetic values only. They reproduce the
official template shape and known validation cases without copying organizer
metadata, credentials, or personal data into Git.
"""

from __future__ import annotations

import hashlib
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "tests" / "fixtures" / "xlsx"
MANIFEST = ROOT / "tests" / "fixtures" / "xlsx_manifest.json"
RULES = ROOT / "config" / "validation_rules.json"
SYNTHETIC_MARKER = "DỮ LIỆU TỔNG HỢP - CHỈ DÙNG KIỂM THỬ"

REPORT_FILENAMES = (
    "BC_T01_Thôn_Phú_Hòa_1.xlsx",
    "BC_T02_Thôn_Phú_Hòa_2.xlsx",
    "BC_T03_Thôn_Thạch_Nham_Đông.xlsx",
    "BC_T04_Thôn_Phước_Thái.xlsx",
    "BC_T05_Thôn_Thái_Lai.xlsx",
    "BC_T06_Thôn_Hòa_Khương_Đông.xlsx",
    "BC_T07_Thôn_Hòa_Khương_Tây.xlsx",
    "BC_T08_Thôn_Phước_Thuận_-_Phước_Hậu.xlsx",
    "BC_T09_Thôn_Phước_Hưng.xlsx",
    "BC_T10_Thôn_Phước_Hưng_Nam.xlsx",
    "BC_T12_Thôn_Trước_Đông.xlsx",
    "BC_T13_Thôn_Diêu_Phong.xlsx",
    "BC_T15_Thôn_Mỹ_Sơn.xlsx",
    "BC_T16_Thôn_Năm.xlsx",
    "BC_T17_Thôn_Hòa_Trung.xlsx",
    "BC_T18_Thôn_Một.xlsx",
    "BC_T19_Thôn_Trung_Nghĩa.xlsx",
    "BC_T20_Thôn_Đông_Sơn.xlsx",
    "BC_T22_Thôn_An_Sơn.xlsx",
)

BASE_VALUES: dict[str, Any] = {
    "CT01": 500,
    "CT02": 1800,
    "CT03": 20,
    "CT04": 15,
    "CT05": 10,
    "CT06": 25,
    "CT07": 300,
    "CT08": 5,
    "CT09": 450,
    "CT10": 1000,
    "CT11": 1700,
    "CT12": 5,
    "CT13": 200,
    "CT14": 0,
}

INTENTIONAL_CASES: dict[str, dict[str, Any]] = {
    "BC_T04_Thôn_Phước_Thái.xlsx": {"CT04": None},
    "BC_T07_Thôn_Hòa_Khương_Tây.xlsx": {"CT07": "ba trăm"},
    "BC_T09_Thôn_Phước_Hưng.xlsx": {"CT02": "2.450"},
    "BC_T12_Thôn_Trước_Đông.xlsx": {"CT02": 25_000},
    "BC_T15_Thôn_Mỹ_Sơn.xlsx": {"CT03": 600},
    "BC_T18_Thôn_Một.xlsx": {"CT11": 1900},
}


def _indicator_rules() -> list[dict[str, Any]]:
    payload = json.loads(RULES.read_text(encoding="utf-8"))
    return [row for row in payload["indicators"] if isinstance(row, dict)]


def _workbook(filename: str, values: dict[str, Any], phone: str) -> Workbook:
    workbook = Workbook()
    workbook.properties.creator = "BaNa SmartLink synthetic test generator"
    workbook.properties.title = SYNTHETIC_MARKER
    worksheet = workbook.active
    worksheet.title = "Phiếu báo cáo"
    worksheet.merge_cells("A1:E1")
    worksheet["A1"] = SYNTHETIC_MARKER
    worksheet["A1"].font = Font(bold=True, color="FFFFFF")
    worksheet["A1"].fill = PatternFill("solid", fgColor="0F5A48")
    worksheet["A1"].alignment = Alignment(horizontal="center")
    worksheet["A3"] = "Tệp kiểm thử"
    worksheet["B3"] = filename.removesuffix(".xlsx")
    worksheet["A9"] = "Người lập biểu"
    worksheet["B9"] = "Cán bộ kiểm thử tổng hợp"
    worksheet["A10"] = "Số điện thoại"
    worksheet["B10"] = phone
    headers = ("Mã CT", "Tên chỉ tiêu", "Đơn vị tính", "Số liệu", "Ghi chú")
    for column, header in enumerate(headers, start=1):
        worksheet.cell(row=13, column=column, value=header)

    for row, rule in enumerate(_indicator_rules(), start=14):
        code = str(rule["code"])
        row_values = (
            code,
            rule["name"],
            rule.get("unit", ""),
            values.get(code),
            "Ca lỗi tổng hợp có chủ đích" if code in INTENTIONAL_CASES.get(filename, {}) else None,
        )
        for column, value in enumerate(row_values, start=1):
            worksheet.cell(row=row, column=column, value=value)

    worksheet.freeze_panes = "A14"
    worksheet.column_dimensions["A"].width = 12
    worksheet.column_dimensions["B"].width = 56
    worksheet.column_dimensions["C"].width = 18
    worksheet.column_dimensions["D"].width = 16
    worksheet.column_dimensions["E"].width = 34
    return workbook


def generate() -> None:
    OUTPUT.mkdir(parents=True, exist_ok=True)
    for stale_file in OUTPUT.glob("*.xlsx"):
        stale_file.unlink()

    files = ("00_BIEU_MAU_TRONG_xa_gui_thon.xlsx", *REPORT_FILENAMES, "TONG_HOP_va_THEO_DOI_TIEN_DO.xlsx")
    for filename in files:
        values = BASE_VALUES.copy()
        values.update(INTENTIONAL_CASES.get(filename, {}))
        if filename == "00_BIEU_MAU_TRONG_xa_gui_thon.xlsx":
            values = {code: None for code in BASE_VALUES}
        phone = "01234" if filename == "BC_T20_Thôn_Đông_Sơn.xlsx" else "0000000000"
        _workbook(filename, values, phone).save(OUTPUT / filename)

    manifest_files = []
    for path in sorted(OUTPUT.glob("*.xlsx")):
        manifest_files.append(
            {
                "name": path.name,
                "sha256": hashlib.sha256(path.read_bytes()).hexdigest(),
                "size": path.stat().st_size,
            }
        )
    MANIFEST.write_text(
        json.dumps(
            {
                "version": 1,
                "classification": "synthetic-test-data",
                "contains_pii": False,
                "files": manifest_files,
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )


if __name__ == "__main__":
    generate()
