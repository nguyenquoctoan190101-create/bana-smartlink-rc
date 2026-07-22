from datetime import datetime
from io import BytesIO
from typing import Any
import unicodedata

from openpyxl import Workbook
from openpyxl.worksheet.page import PageMargins
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Standard styling constants
BLUE_HEADER_FILL = PatternFill(start_color="3B74B4", end_color="3B74B4", fill_type="solid")
WHITE_BOLD_FONT = Font(name="Times New Roman", size=11, bold=True, color="FFFFFF")
BLACK_BOLD_FONT = Font(name="Times New Roman", size=11, bold=True, color="000000")
NORMAL_FONT = Font(name="Times New Roman", size=11, color="000000")
TITLE_FONT = Font(name="Times New Roman", size=13, bold=True, color="000000")
NATION_HEADER_FONT = Font(name="Times New Roman", size=11, bold=True, color="000000")
NATION_SUBHEADER_FONT = Font(name="Times New Roman", size=11, bold=True, italic=True, color="000000")

THIN_BORDER = Border(
    left=Side(style='thin'), 
    right=Side(style='thin'), 
    top=Side(style='thin'), 
    bottom=Side(style='thin')
)

CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
INDICATORS_DICT = {
    "CT01": ("Tổng số hộ dân", "Hộ"),
    "CT02": ("Tổng số nhân khẩu", "Người"),
    "CT03": ("Số hộ nghèo", "Hộ"),
    "CT04": ("Số hộ cận nghèo", "Hộ"),
    "CT05": ("Số người có công với cách mạng đang được quản lý", "Người"),
    "CT06": ("Số đối tượng bảo trợ xã hội đang hưởng trợ cấp", "Người"),
    "CT07": ("Số trẻ em dưới 16 tuổi", "Người"),
    "CT08": ("Số trẻ em có hoàn cảnh đặc biệt", "Người"),
    "CT09": ("Số hộ đạt 'Gia đình văn hóa'", "Hộ"),
    "CT10": ("Số người trong độ tuổi lao động", "Người"),
    "CT11": ("Số người tham gia BHYT", "Người"),
    "CT12": ("Số thành viên Tổ công nghệ số cộng đồng", "Người"),
    "CT13": ("Số người dân được hướng dẫn dùng DVC trực tuyến trong kỳ", "Người"),
    "CT14": ("Số vụ bạo lực gia đình ghi nhận trong kỳ", "Vụ")
}
INDICATOR_CODES = tuple(INDICATORS_DICT)
CURRENT_VILLAGE_ORDER = (
    "Thôn Thạch Nham Đông",
    "Thôn Thạch Nham Tây",
    "Thôn Phước Hưng",
    "Thôn Phú Hòa",
    "Thôn Thái Lai",
    "Thôn Phước Khương",
    "Thôn Hòa Nhơn",
    "Thôn Sơn Phước",
    "Thôn Hòa Ninh",
    "Thôn An Sơn",
)
INDICATOR_RULES = {
    "CT01": "Số nguyên ≥ 0",
    "CT02": "Số nguyên ≥ 0; cảnh báo nếu không xấp xỉ 3–4,5 lần CT01",
    "CT03": "Số nguyên ≥ 0; CT03 ≤ CT01",
    "CT04": "Số nguyên ≥ 0; CT03 + CT04 ≤ CT01",
    "CT05": "Số nguyên ≥ 0",
    "CT06": "Số nguyên ≥ 0",
    "CT07": "Số nguyên ≥ 0; CT07 ≤ CT02",
    "CT08": "Số nguyên ≥ 0; CT08 ≤ CT07",
    "CT09": "Số nguyên ≥ 0; CT09 ≤ CT01",
    "CT10": "Số nguyên ≥ 0; CT10 ≤ CT02",
    "CT11": "Số nguyên ≥ 0; CT11 ≤ CT02",
    "CT12": "Số nguyên ≥ 0; đối chiếu danh sách Tổ CNSCĐ",
    "CT13": "Số nguyên ≥ 0",
    "CT14": "Số nguyên ≥ 0; dữ liệu nội bộ, không công khai",
}


def _safe_excel_text(value: object | None) -> str | None:
    """Prevent untrusted text from being interpreted as a spreadsheet formula."""
    if value is None:
        return None
    text = str(value)
    return f"'{text}" if text.lstrip().startswith(("=", "+", "-", "@")) else text


def _safe_excel_value(value: object | None) -> int | float | str | None:
    """Keep numeric cells typed, null cells blank, and text cells formula-safe."""
    if value is None:
        return None
    if isinstance(value, bool):
        return _safe_excel_text(value)
    if isinstance(value, (int, float)):
        return value
    return _safe_excel_text(value)


def _configure_print_layout(
    worksheet,
    *,
    print_area: str,
    landscape: bool,
    paper_size: int | str,
    title_rows: str | None = None,
) -> None:
    """Keep every exported sheet on one readable printed page.

    The explicit print area prevents Excel from including stray formatted
    cells, while fit-to-page avoids splitting a logical table horizontally.
    Wide 17-column summaries use A3; compact sheets stay on A4.
    """
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.orientation = (
        worksheet.ORIENTATION_LANDSCAPE
        if landscape
        else worksheet.ORIENTATION_PORTRAIT
    )
    worksheet.page_setup.paperSize = paper_size
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 1
    worksheet.print_area = print_area
    if title_rows:
        worksheet.print_title_rows = title_rows
    worksheet.print_options.horizontalCentered = True
    worksheet.page_margins = PageMargins(
        left=0.25,
        right=0.25,
        top=0.5,
        bottom=0.5,
        header=0.2,
        footer=0.2,
    )


def _format_submitted_at(value: object | None) -> str:
    if value is None:
        return ""
    text = str(value)
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).strftime("%d/%m/%Y %H:%M")
    except ValueError:
        return text


def _normalized_sort_text(value: object | None) -> str:
    normalized = unicodedata.normalize("NFD", str(value or "").casefold())
    return "".join(character for character in normalized if not unicodedata.combining(character))


def _ordered_village_items(villages_map: dict) -> list[tuple[Any, Any]]:
    official_positions = {
        _normalized_sort_text(name): index for index, name in enumerate(CURRENT_VILLAGE_ORDER)
    }
    return sorted(
        villages_map.items(),
        key=lambda item: (
            official_positions.get(_normalized_sort_text(item[1]), len(official_positions)),
            _normalized_sort_text(item[1]),
        ),
    )


def _source_label(value: object | None) -> str:
    return {
        "manual": "Nhập tay",
        "excel": "Excel",
        "photo_ocr": "Ảnh/OCR",
        "direct_api": "API trực tiếp",
    }.get(str(value or ""), "Không rõ")


def _workflow_label(value: object | None) -> str:
    return {
        "draft": "Bản nháp",
        "submitted": "Đã nộp",
        "needs_revision": "Cần chỉnh sửa",
        "approved": "Đã duyệt",
        "locked": "Đã khóa",
    }.get(str(value or ""), "Không rõ")


def _timeliness_label(value: object | None) -> str:
    return {
        "not_submitted": "Chưa nộp",
        "on_time": "Đúng hạn",
        "late": "Trễ hạn",
    }.get(str(value or ""), "Chưa xác định")


def _submission_status_label(report: dict[str, Any]) -> str:
    workflow = str(report.get("workflow_status") or "")
    timeliness = report.get("timeliness_status")
    if workflow == "needs_revision":
        suffix = {
            "on_time": "Đúng hạn",
            "late": "Trễ hạn",
        }.get(str(timeliness), "Đã nộp")
        return f"Cần chỉnh sửa · {suffix}"
    if timeliness == "on_time":
        return "Đúng hạn"
    if timeliness == "late":
        return "Trễ hạn"
    return {
        "draft": "Bản nháp",
        "submitted": "Đã nộp",
        "needs_revision": "Cần chỉnh sửa",
        "approved": "Đã duyệt",
        "locked": "Đã khóa",
    }.get(workflow, "Đã nộp")

def generate_village_xlsx_file(period_name: str, report_data: dict, village_name: str) -> bytes:
    """Generates the Single Village Report exactly matching Page 1 format."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Phiếu báo cáo"

    # Header section
    ws.merge_cells('A1:E1')
    ws['A1'] = "CỘNG HÒA XÃ HỘI CHỦ NGHĨA VIỆT NAM"
    ws['A1'].font = NATION_HEADER_FONT
    ws['A1'].alignment = CENTER_ALIGN

    ws.merge_cells('A2:E2')
    ws['A2'] = "Độc lập - Tự do - Hạnh phúc"
    ws['A2'].font = NATION_SUBHEADER_FONT
    ws['A2'].alignment = CENTER_ALIGN

    ws.merge_cells('A4:E4')
    ws['A4'] = "PHIẾU BÁO CÁO SỐ LIỆU VĂN HÓA – XÃ HỘI ĐỊNH KỲ"
    ws['A4'].font = TITLE_FONT
    ws['A4'].alignment = CENTER_ALIGN

    ws.merge_cells('A5:E5')
    ws['A5'] = _safe_excel_text(f"Kỳ báo cáo: {period_name}")
    ws['A5'].font = Font(name="Times New Roman", size=11, italic=True)
    ws['A5'].alignment = CENTER_ALIGN

    # Metadata
    submitter_name = report_data.get("reporter_name") or ""
    submitter_phone = report_data.get("reporter_phone") or ""
    submitted_at = _format_submitted_at(report_data.get("submitted_at"))
        
    ws['A7'] = _safe_excel_text(f"Đơn vị báo cáo (Thôn): {village_name}")
    ws['A8'] = _safe_excel_text(f"Người lập báo cáo: {submitter_name}")
    ws['A9'] = "Chức danh: Cán bộ thôn"
    ws['A10'] = _safe_excel_text(f"Số điện thoại: {submitter_phone}")
    ws['A11'] = _safe_excel_text(f"Thời điểm nộp: {submitted_at}")

    for r in range(7, 12):
        ws[f'A{r}'].font = NORMAL_FONT

    # Table Header
    headers = ["Mã CT", "Tên chỉ tiêu", "Đơn vị tính", "Số liệu", "Ghi chú"]
    ws.append([]) # A12 empty
    ws.append(headers) # row 13
    
    for col_num in range(1, 6):
        cell = ws.cell(row=13, column=col_num)
        cell.fill = BLUE_HEADER_FILL
        cell.font = WHITE_BOLD_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    # Table Data
    vals = report_data.get("values", {})
    for code, (name, unit) in INDICATORS_DICT.items():
        row_data = [
            code,
            name,
            unit,
            _safe_excel_value(vals.get(code)),
            ""
        ]
        ws.append(row_data)
        current_row = ws._current_row
        
        ws.cell(row=current_row, column=1).alignment = CENTER_ALIGN
        ws.cell(row=current_row, column=2).alignment = LEFT_ALIGN
        ws.cell(row=current_row, column=3).alignment = CENTER_ALIGN
        ws.cell(row=current_row, column=4).alignment = CENTER_ALIGN
        ws.cell(row=current_row, column=5).alignment = LEFT_ALIGN
        
        for col_num in range(1, 6):
            cell = ws.cell(row=current_row, column=col_num)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER

    # Set column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 25

    _configure_print_layout(
        ws,
        print_area=f"A1:E{ws.max_row}",
        landscape=False,
        paper_size=ws.PAPERSIZE_A4,
        title_rows="1:12",
    )

    # Export to bytes
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def generate_summary_xlsx_file(period_name: str, reports_data: list, villages_map: dict) -> bytes:
    """Generates the All Villages Summary matching Pages 2,3,4."""
    wb = Workbook()
    
    # Sheet 1: BẢNG TỔNG HỢP SỐ LIỆU
    ws1 = wb.active
    ws1.title = "Bảng tổng hợp"
    
    ws1.merge_cells('A1:Q1')
    ws1['A1'] = "BẢNG TỔNG HỢP SỐ LIỆU VĂN HÓA – XÃ HỘI THEO THÔN"
    ws1['A1'].font = TITLE_FONT
    ws1['A1'].alignment = LEFT_ALIGN
    
    ws1.merge_cells('A2:Q2')
    ws1['A2'] = _safe_excel_text(f"Xã Bà Nà — Kỳ báo cáo: {period_name}")
    ws1['A2'].font = Font(name="Times New Roman", size=11, italic=True)
    ws1['A2'].alignment = LEFT_ALIGN
    
    # Headers
    headers = ["STT", "Thôn"]
    for code, (name, _) in INDICATORS_DICT.items():
        headers.append(f"{code}\n{name}")
    headers.append("Trạng thái nộp")
    
    ws1.append([])
    ws1.append(headers)
    
    # Style Header
    for col_num in range(1, 18):
        cell = ws1.cell(row=4, column=col_num)
        cell.fill = BLUE_HEADER_FILL
        cell.font = WHITE_BOLD_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        
    ws1.column_dimensions['A'].width = 5
    ws1.column_dimensions['B'].width = 25
    for c in range(3, 17):
        ws1.column_dimensions[get_column_letter(c)].width = 14
    ws1.column_dimensions['Q'].width = 15

    # Sort villages
    sorted_villages = _ordered_village_items(villages_map)
    
    # Data Rows
    sums = {code: 0 for code in INDICATORS_DICT.keys()}
    incomplete_totals = {code: False for code in INDICATORS_DICT.keys()}
    submitted_count = 0
    row_idx = 5
    
    for i, (v_id, v_name) in enumerate(sorted_villages, 1):
        report = next((r for r in reports_data if r["village_id"] == v_id), None)
        if report:
            vals = report.get("values", {})
            status_text = _submission_status_label(report)
            
            row_data = [i, _safe_excel_text(v_name)]
            for code in INDICATORS_DICT.keys():
                val = vals.get(code)
                row_data.append(_safe_excel_value(val))
                if isinstance(val, bool) or not isinstance(val, (int, float)):
                    incomplete_totals[code] = True
                else:
                    sums[code] += val
            row_data.append(status_text)
            submitted_count += 1
        else:
            row_data = [i, _safe_excel_text(v_name)]
            for _ in INDICATORS_DICT.keys():
                row_data.append("")
            row_data.append("Chưa nộp")
            
        ws1.append(row_data)
        
        for col_num in range(1, 18):
            cell = ws1.cell(row=row_idx, column=col_num)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            if col_num > 2 and col_num < 17:
                cell.alignment = CENTER_ALIGN
        
        row_idx += 1
        
    # Total Row
    total_row = ["", "TỔNG CỘNG (các thôn đã nộp)"]
    for code in INDICATORS_DICT.keys():
        total_row.append("" if incomplete_totals[code] else sums[code])
    total_row.append(f"{submitted_count}/{len(villages_map)}")
    
    ws1.append(total_row)
    for col_num in range(1, 18):
        cell = ws1.cell(row=row_idx, column=col_num)
        cell.font = BLACK_BOLD_FONT
        cell.border = THIN_BORDER
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        if col_num > 2:
            cell.alignment = CENTER_ALIGN
            
    ws1.freeze_panes = 'C5'
    ws1.row_dimensions[4].height = 72
    _configure_print_layout(
        ws1,
        print_area=f"A1:Q{ws1.max_row}",
        landscape=True,
        paper_size=ws1.PAPERSIZE_A3,
        title_rows="1:4",
    )

    # Sheet 2: THEO DÕI TIẾN ĐỘ
    ws2 = wb.create_sheet("Theo dõi tiến độ")
    ws2.merge_cells('A1:G1')
    ws2['A1'] = "THEO DÕI TIẾN ĐỘ NỘP BÁO CÁO"
    ws2['A1'].font = TITLE_FONT
    ws2['A1'].alignment = LEFT_ALIGN
    
    ws2.merge_cells('A2:G2')
    ws2['A2'] = _safe_excel_text(f"Kỳ báo cáo: {period_name} — Thời điểm tổng hợp: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    ws2['A2'].font = Font(name="Times New Roman", size=11, italic=True)
    ws2['A2'].alignment = LEFT_ALIGN
    
    ws2.append([])
    headers_progress = ["STT", "Thôn", "Người lập", "SĐT", "Thời điểm nộp", "Trạng thái", "Số ngày trễ"]
    ws2.append(headers_progress)
    
    for col_num in range(1, 8):
        cell = ws2.cell(row=4, column=col_num)
        cell.fill = BLUE_HEADER_FILL
        cell.font = WHITE_BOLD_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        
    row_idx = 5
    for i, (v_id, v_name) in enumerate(sorted_villages, 1):
        report = next((r for r in reports_data if r["village_id"] == v_id), None)
        if report:
            submitter_name = report.get("reporter_name") or ""
            submitter_phone = report.get("reporter_phone") or ""
            submitted_at = _format_submitted_at(report.get("submitted_at"))
            status_text = _submission_status_label(report)
            days_late = _safe_excel_value(report.get("days_late"))
        else:
            submitter_name = ""
            submitter_phone = ""
            submitted_at = ""
            status_text = "Chưa nộp"
            days_late = ""
            
        row_data = [
            i,
            _safe_excel_text(v_name),
            _safe_excel_text(submitter_name),
            _safe_excel_text(submitter_phone),
            _safe_excel_text(submitted_at),
            status_text,
            days_late,
        ]
        ws2.append(row_data)
        
        for col_num in range(1, 8):
            cell = ws2.cell(row=row_idx, column=col_num)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            if col_num in [1, 4, 5, 6, 7]:
                cell.alignment = CENTER_ALIGN
                
            if status_text == "Chưa nộp" and col_num == 6:
                cell.fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
            elif status_text == "Trễ hạn" and col_num == 6:
                cell.fill = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid")
            elif col_num == 6:
                cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
                
        row_idx += 1
        
    ws2.column_dimensions['A'].width = 5
    ws2.column_dimensions['B'].width = 25
    ws2.column_dimensions['C'].width = 25
    ws2.column_dimensions['D'].width = 15
    ws2.column_dimensions['E'].width = 20
    ws2.column_dimensions['F'].width = 20
    ws2.column_dimensions['G'].width = 15
    ws2.row_dimensions[4].height = 34
    _configure_print_layout(
        ws2,
        print_area=f"A1:G{ws2.max_row}",
        landscape=True,
        paper_size=ws2.PAPERSIZE_A4,
        title_rows="1:4",
    )

    # Sheet 2 in the official reference workbook: a compact operational
    # dashboard. It intentionally contains typed values rather than charts so
    # it remains accessible and verifiable in every spreadsheet viewer.
    dashboard = wb.create_sheet("Dashboard")
    wb.move_sheet(dashboard, offset=-1)
    dashboard.merge_cells("A1:B1")
    dashboard["A1"] = "DASHBOARD TỔNG QUAN"
    dashboard["A1"].font = TITLE_FONT
    dashboard.append([])
    dashboard.append(["Chỉ số", "Giá trị"])
    on_time_count = sum(1 for report in reports_data if report.get("timeliness_status") == "on_time")
    late_count = sum(1 for report in reports_data if report.get("timeliness_status") == "late")
    dashboard_rows = [
        ("Tổng số thôn", len(villages_map)),
        ("Đã nộp", submitted_count),
        ("Đúng hạn", on_time_count),
        ("Trễ hạn", late_count),
        ("Chưa nộp", len(villages_map) - submitted_count),
        ("Tỷ lệ nộp (%)", round(submitted_count * 100 / len(villages_map), 1) if villages_map else 0),
        ("", ""),
        ("Chỉ tiêu (tổng trong phạm vi)", "Giá trị"),
        ("Tổng số hộ dân", "—" if incomplete_totals["CT01"] else sums["CT01"]),
        ("Số hộ nghèo", "—" if incomplete_totals["CT03"] else sums["CT03"]),
        ("Số hộ cận nghèo", "—" if incomplete_totals["CT04"] else sums["CT04"]),
        ("Số người có công với cách mạng đang được quản lý", "—" if incomplete_totals["CT05"] else sums["CT05"]),
        ("Đối tượng bảo trợ xã hội", "—" if incomplete_totals["CT06"] else sums["CT06"]),
        ("Hộ đạt Gia đình văn hóa", "—" if incomplete_totals["CT09"] else sums["CT09"]),
    ]
    for row in dashboard_rows:
        dashboard.append([_safe_excel_text(row[0]), _safe_excel_value(row[1])])
    for row in range(3, dashboard.max_row + 1):
        for column in range(1, 3):
            cell = dashboard.cell(row=row, column=column)
            cell.border = THIN_BORDER
            cell.font = BLACK_BOLD_FONT if row in {3, 11} else NORMAL_FONT
            if row in {3, 11}:
                cell.fill = BLUE_HEADER_FILL
                cell.font = WHITE_BOLD_FONT
            if column == 2 and row not in {3, 10, 11}:
                cell.alignment = CENTER_ALIGN
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.0" if row == 9 else "#,##0"
    dashboard.column_dimensions["A"].width = 48
    dashboard.column_dimensions["B"].width = 18
    dashboard.freeze_panes = "A4"
    dashboard.sheet_view.showGridLines = False
    _configure_print_layout(
        dashboard,
        print_area=f"A1:B{dashboard.max_row}",
        landscape=False,
        paper_size=dashboard.PAPERSIZE_A4,
        title_rows="1:3",
    )

    dictionary = wb.create_sheet("Từ điển dữ liệu")
    dictionary.merge_cells("A1:E1")
    dictionary["A1"] = "TỪ ĐIỂN DỮ LIỆU (DATA DICTIONARY)"
    dictionary["A1"].font = TITLE_FONT
    dictionary.append([])
    dictionary.append(["Mã CT", "Tên chỉ tiêu", "Đơn vị tính", "Kiểu dữ liệu", "Ràng buộc / kiểm tra hợp lệ"])
    for code, (name, unit) in INDICATORS_DICT.items():
        dictionary.append([code, name, unit, "Số nguyên", INDICATOR_RULES[code]])
    for column in range(1, 6):
        cell = dictionary.cell(row=3, column=column)
        cell.fill = BLUE_HEADER_FILL
        cell.font = WHITE_BOLD_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    for row in range(4, dictionary.max_row + 1):
        for column in range(1, 6):
            dictionary.cell(row=row, column=column).border = THIN_BORDER
            dictionary.cell(row=row, column=column).font = NORMAL_FONT
            dictionary.cell(row=row, column=column).alignment = LEFT_ALIGN
    for column, width in {"A": 10, "B": 58, "C": 16, "D": 16, "E": 62}.items():
        dictionary.column_dimensions[column].width = width
    dictionary.freeze_panes = "A4"
    dictionary.row_dimensions[3].height = 38
    _configure_print_layout(
        dictionary,
        print_area=f"A1:E{dictionary.max_row}",
        landscape=True,
        paper_size=dictionary.PAPERSIZE_A4,
        title_rows="1:3",
    )

    warnings = wb.create_sheet("Cảnh báo dữ liệu")
    warnings.append(["Thôn", "Mã báo cáo", "Chỉ tiêu", "Loại cảnh báo", "Nội dung", "Trạng thái"])
    warning_count = 0
    for report in reports_data:
        village_name = villages_map.get(str(report.get("village_id")), str(report.get("village_id") or ""))
        for flag in report.get("validation_flags") or []:
            warnings.append([
                _safe_excel_text(village_name),
                _safe_excel_text(report.get("id")),
                _safe_excel_text(flag.get("ct_code")),
                _safe_excel_text(flag.get("error_type")),
                _safe_excel_text(flag.get("message")),
                "Đã xử lý" if flag.get("resolved") else "Cần xem",
            ])
            warning_count += 1
    if warning_count == 0:
        warnings.append(["Không có cảnh báo chưa xử lý trong lát cắt xuất bản", "", "", "", "", ""])
        warnings.merge_cells("A2:F2")
        warnings["A2"].alignment = CENTER_ALIGN
    for column in range(1, 7):
        cell = warnings.cell(row=1, column=column)
        cell.fill = BLUE_HEADER_FILL
        cell.font = WHITE_BOLD_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    for row in range(2, warnings.max_row + 1):
        for column in range(1, 7):
            warnings.cell(row=row, column=column).border = THIN_BORDER
            warnings.cell(row=row, column=column).font = NORMAL_FONT
    for column, width in {"A": 25, "B": 38, "C": 12, "D": 20, "E": 62, "F": 16}.items():
        warnings.column_dimensions[column].width = width
    warnings.freeze_panes = "A2"
    warnings.row_dimensions[1].height = 34
    _configure_print_layout(
        warnings,
        print_area=f"A1:F{warnings.max_row}",
        landscape=True,
        paper_size=warnings.PAPERSIZE_A4,
        title_rows="1:1",
    )

    sources = wb.create_sheet("Nguồn dữ liệu")
    sources.append([
        "Thôn", "Mã báo cáo", "Nguồn nhập", "Phiên bản báo cáo", "Bộ quy tắc",
        "Trạng thái nghiệp vụ", "Trạng thái đúng hạn", "Thời điểm nộp",
    ])
    village_positions = {
        str(village_id): index
        for index, (village_id, _name) in enumerate(_ordered_village_items(villages_map))
    }
    for report in sorted(
        reports_data,
        key=lambda item: (
            village_positions.get(str(item.get("village_id")), len(village_positions)),
            str(item.get("village_id") or ""),
        ),
    ):
        sources.append([
            _safe_excel_text(villages_map.get(str(report.get("village_id")), str(report.get("village_id") or ""))),
            _safe_excel_text(report.get("id")),
            _safe_excel_text(_source_label(report.get("report_source"))),
            _safe_excel_value(report.get("version")),
            _safe_excel_text(report.get("rule_version") or "2026-07"),
            _safe_excel_text(_workflow_label(report.get("workflow_status"))),
            _safe_excel_text(_timeliness_label(report.get("timeliness_status"))),
            _safe_excel_text(_format_submitted_at(report.get("submitted_at"))),
        ])
    for column in range(1, 9):
        cell = sources.cell(row=1, column=column)
        cell.fill = BLUE_HEADER_FILL
        cell.font = WHITE_BOLD_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    for row in range(2, sources.max_row + 1):
        for column in range(1, 9):
            sources.cell(row=row, column=column).border = THIN_BORDER
            sources.cell(row=row, column=column).font = NORMAL_FONT
    for column, width in {"A": 25, "B": 38, "C": 18, "D": 18, "E": 16, "F": 22, "G": 20, "H": 22}.items():
        sources.column_dimensions[column].width = width
    sources.freeze_panes = "A2"
    sources.row_dimensions[1].height = 42
    _configure_print_layout(
        sources,
        print_area=f"A1:H{sources.max_row}",
        landscape=True,
        paper_size=sources.PAPERSIZE_A4,
        title_rows="1:1",
    )

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
