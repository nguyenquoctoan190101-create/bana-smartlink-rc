from __future__ import annotations

import json
import unicodedata
from io import BytesIO
from pathlib import Path
from typing import Any, TypedDict
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from services.validator import coerce_storage_value


RULES_PATH = Path(__file__).resolve().parents[1] / "config" / "validation_rules.json"
REPORT_SHEET_NAME = "Phiếu báo cáo"
FIRST_INDICATOR_ROW = 13
EXCEL_EXTRACTOR_VERSION = "1.0"


class ParsedExcelMetadata(TypedDict):
    period_name: str | None
    village_name: str | None
    reporter_name: str | None
    reporter_title: str | None
    reporter_phone: str | None
    deadline: str | None


class ParsedExcelReport(TypedDict):
    values: dict[str, Any]
    notes: dict[str, str | None]
    metadata: ParsedExcelMetadata
    evidence: dict[str, "ParsedExcelFieldEvidence"]


class ParsedExcelFieldEvidence(TypedDict):
    raw_value: int | float | str | None
    normalized_value: int | None
    confidence: float
    source_page: None
    source_region: str
    extractor: str
    method: str
    version: str
    flags: list[str]
    requires_review: bool


class ExcelReportParseError(RuntimeError):
    """Raised when the official Excel report template cannot be parsed."""


def parse_official_report_excel(file_bytes: bytes) -> ParsedExcelReport:
    """Parse the official thon Excel template into report values and notes."""
    expected_codes = _load_indicator_codes()
    try:
        workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
        raise ExcelReportParseError("Không đọc được file Excel.") from exc

    sheet_name = _find_report_sheet(workbook, expected_codes)

    worksheet = workbook[sheet_name]
    values: dict[str, Any] = {code: None for code in expected_codes}
    notes: dict[str, str | None] = {code: None for code in expected_codes}
    evidence: dict[str, ParsedExcelFieldEvidence] = {}
    metadata: ParsedExcelMetadata = {
        "period_name": _clean_prefixed_text(worksheet["A5"].value, "Kỳ báo cáo:"),
        "village_name": _clean_text(worksheet["B7"].value),
        "reporter_name": _clean_text(worksheet["B8"].value),
        "reporter_title": _clean_text(worksheet["B9"].value),
        "reporter_phone": _clean_text(worksheet["B10"].value),
        "deadline": _clean_text(worksheet["B11"].value),
    }

    for row in worksheet.iter_rows(min_row=FIRST_INDICATOR_ROW):
        ct_code = _clean_code(row[0].value if len(row) > 0 else None)
        if ct_code not in expected_codes:
            continue

        values[ct_code] = row[3].value if len(row) > 3 else None
        notes[ct_code] = _clean_note(row[4].value if len(row) > 4 else None)
        normalized_value = coerce_storage_value(values[ct_code])
        evidence[ct_code] = {
            "raw_value": values[ct_code],
            "normalized_value": normalized_value,
            "confidence": 1.0 if normalized_value is not None else 0.0,
            "source_page": None,
            "source_region": f"{sheet_name}!D{row[0].row}",
            "extractor": "openpyxl",
            "method": "official_template_cell",
            "version": EXCEL_EXTRACTOR_VERSION,
            "flags": [],
            "requires_review": normalized_value is None,
        }

    for code in expected_codes:
        if code not in evidence:
            evidence[code] = {
                "raw_value": None,
                "normalized_value": None,
                "confidence": 0.0,
                "source_page": None,
                "source_region": f"{sheet_name}!D?",
                "extractor": "openpyxl",
                "method": "official_template_cell",
                "version": EXCEL_EXTRACTOR_VERSION,
                "flags": ["SOURCE_CELL_NOT_FOUND"],
                "requires_review": True,
            }

    return {
        "values": values,
        "notes": notes,
        "metadata": metadata,
        "evidence": evidence,
    }


def _normalize_sheet_name(value: Any) -> str:
    """Match the official sheet despite harmless whitespace/Unicode differences."""
    text = unicodedata.normalize("NFKC", str(value or ""))
    return " ".join(text.split()).casefold()


def _sheet_name_key(value: Any) -> str:
    """Ignore harmless accents, casing and punctuation in copied template tabs."""
    text = unicodedata.normalize("NFKD", str(value or ""))
    without_accents = "".join(char for char in text if not unicodedata.combining(char))
    return "".join(char for char in without_accents.casefold() if char.isalnum())


def _find_report_sheet(workbook: Any, expected_codes: set[str]) -> str:
    """Locate the report tab by its canonical name, then a unique CT layout."""
    exact_name = _normalize_sheet_name(REPORT_SHEET_NAME)
    for name in workbook.sheetnames:
        if _normalize_sheet_name(name) == exact_name:
            return name

    expected_key = _sheet_name_key(REPORT_SHEET_NAME)
    for name in workbook.sheetnames:
        if _sheet_name_key(name) == expected_key:
            return name

    candidates: list[str] = []
    for name in workbook.sheetnames:
        worksheet = workbook[name]
        found_codes = {
            _clean_code(row[0].value if row else None)
            for row in worksheet.iter_rows(min_row=1, max_col=1)
        }
        # A single CT code is not enough: only accept a sheet with most of the
        # official structure, avoiding accidental imports from unrelated tabs.
        if len(found_codes & expected_codes) >= min(8, len(expected_codes)):
            candidates.append(name)

    if len(candidates) == 1:
        return candidates[0]
    if len(candidates) > 1:
        raise ExcelReportParseError(
            "Có nhiều sheet có dữ liệu CT01–CT14; hãy đổi tên sheet báo cáo thành 'Phiếu báo cáo'."
        )
    raise ExcelReportParseError(
        "Không tìm thấy sheet báo cáo. Dùng tên 'Phiếu báo cáo' hoặc một sheet chứa ít nhất 8 chỉ tiêu CT01–CT14."
    )


def _load_indicator_codes() -> set[str]:
    with RULES_PATH.open("r", encoding="utf-8") as rules_file:
        payload = json.load(rules_file)

    indicators = payload.get("indicators", [])
    if not isinstance(indicators, list):
        raise ValueError("validation_rules.json must contain an indicators list")

    return {
        str(indicator["code"]).strip().upper()
        for indicator in indicators
        if isinstance(indicator, dict) and indicator.get("code")
    }


def _clean_code(value: Any) -> str:
    if value is None:
        return ""

    return str(value).strip().upper()


def _clean_note(value: Any) -> str | None:
    if value is None:
        return None

    note = str(value).strip()
    return note or None


def _clean_text(value: Any) -> str | None:
    if value is None:
        return None

    text = str(value).strip()
    return text or None


def _clean_prefixed_text(value: Any, prefix: str) -> str | None:
    text = _clean_text(value)
    if text is None:
        return None
    if text.casefold().startswith(prefix.casefold()):
        text = text[len(prefix):].strip()
    return text or None


__all__ = [
    "EXCEL_EXTRACTOR_VERSION",
    "ExcelReportParseError",
    "parse_official_report_excel",
]
