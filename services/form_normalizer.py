from __future__ import annotations

import json
import re
import unicodedata
from collections.abc import Mapping
from io import BytesIO
from pathlib import Path
from typing import Any, TypedDict
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

try:
    from thefuzz import fuzz
except ImportError:  # pragma: no cover - keeps local syntax checks lightweight.
    fuzz = None  # type: ignore[assignment]


RULES_PATH = Path(__file__).resolve().parents[1] / "config" / "validation_rules.json"
SYNONYMS_JSON_PATH = (
    Path(__file__).resolve().parents[1] / "config" / "field_synonyms.json"
)
AUTO_MATCH_THRESHOLD = 85
REVIEW_MATCH_THRESHOLD = 60
MAX_SCAN_ROWS = 300
MAX_SCAN_COLUMNS = 60
VALUE_LOOKAHEAD = 6
VALID_CT_CODE_RE = re.compile(r"^CT(?:0[1-9]|1[0-4])$")


class NormalizedIndicator(TypedDict):
    value: int | str | None
    confidence: float
    matched_from: str | None
    status: str
    message: str
    requires_confirmation: bool


class FormNormalizationError(RuntimeError):
    """Raised when an Excel workbook cannot be normalized."""


class _IndicatorRule(TypedDict):
    code: str
    name: str


class _Candidate(TypedDict):
    label: str
    value: Any
    sheet_name: str
    cell: str


def normalize_field_name(value: str) -> str:
    """Return the stable key shared by file and database mappings."""
    return _normalize_text(value)


def _validated_synonyms(payload: object) -> dict[str, str]:
    if not isinstance(payload, dict):
        raise FormNormalizationError("Cấu hình ánh xạ trường không hợp lệ.")

    validated: dict[str, str] = {}
    for original_name, ct_code in payload.items():
        if not isinstance(original_name, str) or not isinstance(ct_code, str):
            raise FormNormalizationError("Cấu hình ánh xạ trường không hợp lệ.")
        normalized_name = normalize_field_name(original_name)
        normalized_code = ct_code.strip().upper()
        if not normalized_name or VALID_CT_CODE_RE.fullmatch(normalized_code) is None:
            raise FormNormalizationError("Cấu hình ánh xạ trường không hợp lệ.")
        validated[normalized_name] = normalized_code
    return validated


def load_synonyms(path: Path | None = None) -> dict[str, str]:
    """Load the optional bundled fallback mapping with strict validation."""
    target = path or SYNONYMS_JSON_PATH
    if not target.exists():
        return {}
    try:
        with target.open("r", encoding="utf-8") as file_handle:
            payload = json.load(file_handle)
    except (OSError, UnicodeError, json.JSONDecodeError) as exc:
        raise FormNormalizationError("Không thể đọc cấu hình ánh xạ trường.") from exc
    return _validated_synonyms(payload)


def save_synonym(
    original_name: str,
    ct_code: str,
    path: Path | None = None,
) -> None:
    """Persist a fallback mapping atomically.

    Runtime mappings are stored by database RPC. This helper remains available
    for fixtures and offline packaging without mutating source during tests.
    """
    normalized_name = normalize_field_name(original_name)
    normalized_code = ct_code.strip().upper()
    if not normalized_name or VALID_CT_CODE_RE.fullmatch(normalized_code) is None:
        raise FormNormalizationError("Tên trường hoặc mã chỉ tiêu không hợp lệ.")

    target = path or SYNONYMS_JSON_PATH
    synonyms = load_synonyms(target)
    synonyms[normalized_name] = normalized_code
    temporary_target = target.with_suffix(f"{target.suffix}.tmp")
    try:
        target.parent.mkdir(parents=True, exist_ok=True)
        with temporary_target.open("w", encoding="utf-8", newline="\n") as file_handle:
            json.dump(synonyms, file_handle, ensure_ascii=False, indent=2)
            file_handle.write("\n")
        temporary_target.replace(target)
    except (OSError, UnicodeError, TypeError, ValueError) as exc:
        temporary_target.unlink(missing_ok=True)
        raise FormNormalizationError("Không thể lưu cấu hình ánh xạ trường.") from exc


def normalize_excel(
    file_bytes: bytes,
    synonyms: Mapping[str, str] | None = None,
) -> dict[str, NormalizedIndicator]:
    """Normalize arbitrary Excel forms into CT01-CT14 before validation."""
    rules = _load_indicator_rules()
    candidates = _extract_candidates(file_bytes)
    normalized: dict[str, NormalizedIndicator] = {}
    merged_synonyms = load_synonyms()
    if synonyms is not None:
        merged_synonyms.update(_validated_synonyms(dict(synonyms)))

    for rule in rules:
        code = rule["code"]
        best_candidate: _Candidate | None = None
        best_score = 0

        for candidate in candidates:
            score = _score_label(candidate["label"], rule, merged_synonyms)
            if score > best_score:
                best_candidate = candidate
                best_score = score

        if best_candidate is None or best_score < REVIEW_MATCH_THRESHOLD:
            normalized[code] = _unidentified_indicator()
            continue

        confidence = round(best_score / 100, 2)
        matched_from = best_candidate["label"]
        value = _normalize_value(best_candidate["value"])
        requires_confirmation = best_score < AUTO_MATCH_THRESHOLD

        normalized[code] = {
            "value": value,
            "confidence": confidence,
            "matched_from": matched_from,
            "status": (
                "needs_confirmation"
                if requires_confirmation
                else "auto_mapped"
            ),
            "message": (
                f"Hệ thống gợi ý đây là {code} — cần xác nhận."
                if requires_confirmation
                else "Tự động ánh xạ từ mẫu Excel."
            ),
            "requires_confirmation": requires_confirmation,
        }

    return normalized


def _load_indicator_rules() -> list[_IndicatorRule]:
    with RULES_PATH.open("r", encoding="utf-8") as rules_file:
        payload = json.load(rules_file)

    indicators = payload.get("indicators", [])
    if not isinstance(indicators, list):
        raise ValueError("validation_rules.json must contain an indicators list")

    rules: list[_IndicatorRule] = []
    for indicator in indicators:
        if isinstance(indicator, dict):
            rules.append(
                {
                    "code": str(indicator["code"]),
                    "name": str(indicator["name"]),
                }
            )
    return rules


def _extract_candidates(file_bytes: bytes) -> list[_Candidate]:
    try:
        workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
        raise FormNormalizationError("Không đọc được file Excel.") from exc

    candidates: list[_Candidate] = []
    for worksheet in workbook.worksheets:
        rows = list(
            worksheet.iter_rows(
                min_row=1,
                max_row=min(worksheet.max_row, MAX_SCAN_ROWS),
                min_col=1,
                max_col=min(worksheet.max_column, MAX_SCAN_COLUMNS),
            )
        )
        candidates.extend(_row_candidates(worksheet.title, rows))
        candidates.extend(_column_candidates(worksheet.title, rows))
    return candidates


def _row_candidates(
    sheet_name: str,
    rows: list[tuple[Any, ...]],
) -> list[_Candidate]:
    candidates: list[_Candidate] = []
    for row in rows:
        for index, cell in enumerate(row):
            label = _text_value(cell.value)
            if label is None:
                continue

            value_cell = _first_value_cell(
                row[index + 1 : index + 1 + VALUE_LOOKAHEAD]
            )
            if value_cell is not None:
                candidates.append(
                    {
                        "label": label,
                        "value": value_cell.value,
                        "sheet_name": sheet_name,
                        "cell": cell.coordinate,
                    }
                )
    return candidates


def _column_candidates(
    sheet_name: str,
    rows: list[tuple[Any, ...]],
) -> list[_Candidate]:
    candidates: list[_Candidate] = []
    if not rows:
        return candidates

    column_count = max(len(row) for row in rows)
    for column_index in range(column_count):
        column = [row[column_index] for row in rows if column_index < len(row)]
        for index, cell in enumerate(column):
            label = _text_value(cell.value)
            if label is None:
                continue

            value_cell = _first_value_cell(
                column[index + 1 : index + 1 + VALUE_LOOKAHEAD]
            )
            if value_cell is not None:
                candidates.append(
                    {
                        "label": label,
                        "value": value_cell.value,
                        "sheet_name": sheet_name,
                        "cell": cell.coordinate,
                    }
                )
    return candidates


def _score_label(
    label: str,
    rule: _IndicatorRule,
    synonyms: dict[str, str],
) -> int:
    normalized_label = _normalize_text(label)
    if synonyms.get(normalized_label) == rule["code"]:
        return 100

    official_name = _normalize_text(rule["name"])
    code = _normalize_text(rule["code"])
    if code and re.search(rf"\b{re.escape(code)}\b", normalized_label):
        return 100

    if fuzz is not None:
        return max(
            fuzz.token_set_ratio(normalized_label, official_name),
            fuzz.partial_ratio(normalized_label, official_name),
        )

    return max(
        int(_difflib_score(normalized_label, official_name) * 100),
        _token_overlap_score(normalized_label, official_name),
    )


def _difflib_score(left: str, right: str) -> float:
    from difflib import SequenceMatcher

    return SequenceMatcher(a=left, b=right).ratio()


def _token_overlap_score(left: str, right: str) -> int:
    left_tokens = set(left.split())
    right_tokens = set(right.split())
    if not left_tokens or not right_tokens:
        return 0
    if left_tokens <= right_tokens or right_tokens <= left_tokens:
        return 100
    common_tokens = left_tokens & right_tokens
    return int((len(common_tokens) / max(len(left_tokens), len(right_tokens))) * 100)


def _normalize_text(value: str) -> str:
    without_accents = "".join(
        character
        for character in unicodedata.normalize("NFD", value)
        if unicodedata.category(character) != "Mn"
    )
    lowered = without_accents.casefold()
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", lowered)).strip()


def _text_value(value: Any) -> str | None:
    if not isinstance(value, str):
        return None
    stripped_value = value.strip()
    if not stripped_value:
        return None
    if re.fullmatch(r"[-+]?\d+(?:[.,]\d+)?", stripped_value):
        return None
    return stripped_value


def _first_value_cell(cells: list[Any]) -> Any | None:
    for cell in cells:
        if _is_value_like(cell.value):
            return cell
    return None


def _is_value_like(value: Any) -> bool:
    if value is None or isinstance(value, bool):
        return False
    if isinstance(value, (int, float)):
        return True
    if isinstance(value, str):
        stripped_value = value.strip()
        return bool(stripped_value) and (
            _text_value(stripped_value) is None
            or re.fullmatch(r"[-+]?\d+\s*\D+", stripped_value) is not None
        )
    return False


def _normalize_value(value: Any) -> int | str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else str(value)
    if not isinstance(value, str):
        return str(value)

    stripped_value = value.strip()
    if not stripped_value:
        return None
    if re.search(r"\d[.,]\d", stripped_value):
        return stripped_value
    if re.fullmatch(r"[-+]?\d+", stripped_value):
        return int(stripped_value)

    unit_value_match = re.fullmatch(r"([-+]?\d+)\s*\D+", stripped_value)
    if unit_value_match is not None:
        return int(unit_value_match.group(1))
    return stripped_value


def _unidentified_indicator() -> NormalizedIndicator:
    return {
        "value": None,
        "confidence": 0.0,
        "matched_from": None,
        "status": "unidentified",
        "message": "Không xác định được chỉ tiêu trong mẫu Excel.",
        "requires_confirmation": True,
    }


__all__ = [
    "FormNormalizationError",
    "load_synonyms",
    "normalize_excel",
    "normalize_field_name",
    "save_synonym",
]
